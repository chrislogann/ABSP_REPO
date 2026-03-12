import os
import openpyxl
import csv
import logging

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s %(levelname)s - %(message)s')
logging.disable()

"""
Script 14_excel_to_csv.py convert excel rows into a csv file.
Each excel workseet will geenerate a file with <workbook>_<worksheet>.csv filename.
Export files output into a folder called excel_csv.
"""

## Return excel filepaths
def get_excel_filepaths(pDirectory):

    """
    Provides a list of excel filepaths from a given directory.

    Params:
    pDirectory: Folder location

    Return:
    vFilePathList: List containing excel filepaths
    """

    """
    Entry Point
    """

    logging.debug(f"Executing function get_excel_filepaths")

    vFilepathList= []
    for path,subfolder,filenames in os.walk(pDirectory):

        for filename in filenames:

            if not str(filename).endswith(".xlsx"):
                continue
            
            filepath = os.path.join(path,filename)
            vFilepathList.append(filepath)
    
    logging.debug(f"Function get_excel_filepaths complete")
    return vFilepathList

## Gather excel content
def return_excel_data(pFilePath):

    """
    Function return_excel_data outputs export data from a given excel workbook filepath.
    An export filename is assigned from pFilePath's basename and active excel worksheet.

    Params:
    pFilePath: Filepath to a given excel workbook.

    Returns:
    vExportDataList: A list containing export data from each excel workbook's filepath
    """

    def get_sheet_data(pActiveSheet):

        """
        Function get_sheet_data gathers data from a given active excel sheet.

        Params:
        pActiveSheet: Input excel sheet to extract data from

        Returns:
        vSheetDataList: A list containing all data from an excel sheet.

        """

        """
        Entry Point
        """

        logging.debug(f"Executing function get_sheet_data")

        vSheetDataList = []
        for row in pActiveSheet.iter_rows(values_only=True):
            vSheetDataList.append(list(row))

        logging.debug(f"Function get_sheet_data complete")
        return vSheetDataList

    """
    Entry Point
    """

    logging.debug(f"Executing function return_excel_data")

    ## Creating export filename
    filename = os.path.basename(pFilePath)
    filename_no_extension = os.path.splitext(filename)[0]
    wb = openpyxl.open(pFilePath)

    ## Getting export data
    vExportDataList = []
    logging.debug(f"Executing wb.sheetnames for loop")
    for sheetname in wb.sheetnames:

        export_filename = f"{filename_no_extension}_{sheetname}.csv"
        active_sheet = wb[sheetname]

        vSheetData = get_sheet_data(active_sheet)
        vExportData = (export_filename,vSheetData)
        vExportDataList.append(vExportData)
    logging.debug(f"wb.sheetnames loop complete")

    logging.debug(f"Function return_excel_data complete")
    return vExportDataList

def gather_excel_content(pFilePathList):
    
    """
    Aquires data from given excel filepaths.

    Params:
    pFilePathList: A list with excel filepaths

    Returns:
    vExcelDataList: A list containing all sheet data for a filepath.

    """

    logging.debug(f"Function gather_excel_content complete")

    vExcelDataList = []

    logging.debug(f"Running pFilePathList loop")
    for filepath in pFilePathList:

        vExcelData = return_excel_data(filepath)
        vExcelDataList.extend(vExcelData)
    logging.debug(f"Loop pFilePathList complete")

    logging.debug(f"Function gather_excel_content complete")
    return vExcelDataList

## Generate csv export
def output_csv_file (pDirectory,pExcelDataList):
    
    """
    Function output_csv_file takes a list of data and outputs a csv file
    to a given directory.

    Params:
    pDirectory: Location to output file
    pExcelDataList: A list containing excel data.

    Returns:
    None

    """

    def generate_output_file(pExportFilepath,pExportDataList):

        """
        Function generate_output_file creates an output csv from
        a given export filepath and list of data.

        Params:
        pExportFilepath: Location to output file
        pExportDataList: A list containing list of data content.

        Returns:
        None
        """

        with open(pExportFilepath, "w", newline="") as csv_file:
            csvWriter = csv.writer(csv_file)
            csvWriter.writerows(pExportDataList)

    """
    Entry Point
    """

    logging.debug(f"Executing function output_csv_file")

    excel_csv_folder = os.path.join(pDirectory,"excel_csv_folder")
    os.makedirs(excel_csv_folder,exist_ok=True)

    logging.debug(f"Executing pExcelDataList loop")
    for export_filename, export_data in pExcelDataList:

        export_filepath = os.path.join(excel_csv_folder, export_filename)
        generate_output_file(export_filepath, export_data)
    logging.debug(f"pExcelDataList loop complete")

    logging.debug(f"Function output_csv_file complete")

def main():
    """
    Function main conducts all predefined functions in 14_excel_to_csv.py.

    Params:
    None

    Returns:
    None
    """
    vDirectory = os.getcwd()
    vFilePathList = get_excel_filepaths(vDirectory)
    vExcelDataList = gather_excel_content(vFilePathList)
    output_csv_file(vDirectory,vExcelDataList)

main()