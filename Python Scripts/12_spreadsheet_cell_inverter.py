import sys
import openpyxl
import re

"""
Script 12_spreadsheet_cell_inverter.py pivots an input excel table.
Outputs a copy of input workbook.
"""

def pivot_data(pWorkbook):
    ## Open Workbook
    workbook_name = pWorkbook
    wb = openpyxl.open(workbook_name)
    sheet = wb.active

    ## Transpose table

    ## Coords (y,x)
    ## Gathering column data
    nested_list = []
    for y in range(1,sheet.max_column+1):

        column_list = []
        for x in range(1,sheet.max_row+1):

            row_data = sheet.cell(row = x,column = y).value
            column_list.append(row_data)

            sheet.cell(row = x,column = y).value = None
        
        nested_list.append(column_list)

    nested_list_len = len(nested_list)

    ## coords (x,y)
    ## Inserting column data as rows
    for x in range(nested_list_len):
        
        list_len = len(nested_list[x])
        for y in range(list_len):
            row_data = sheet.cell(row = x+1,column = y+1).value = nested_list[x][y]


    ## Output copy of workbook
    match = re.match(r"^(.+?)\.(.+?)$",workbook_name)
    workbook_name_no_extension = match.group(1)
    newworkbook = f"{workbook_name_no_extension}_COPY.xlsx"
    wb.save(newworkbook)
    print(f"Generated workbook {newworkbook}")

workbook = "PIVOT_SAMPLE.xlsx"
pivot_data(workbook)