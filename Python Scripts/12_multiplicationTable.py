
import sys
import openpyxl

"""
Takes a command line number and creates an excel multiplication table.
Row one and column A are used for labels.

Called as: python 12_multiplicationTable.py <number>
Outputs excel workbook Multiplication_table.xlsx
"""

# Returning sys.argv
try:
    num = int(sys.argv[1])
except IndexError:
    print("Whole Number argument missing. Quitting program")
    quit()
except TypeError,ValueError:
    print("Argument not a number. Quitting Program")
    quit()

# Creating Excel workbook

wb = openpyxl.Workbook()
wb.sheetnames
sheet = wb.active
sheet.title = 'Multiplication Table'

# Adding Labels
num_list = []
while num != 0:

    num_list.append(num)
    num -= 1

num_list.sort()

for rowNum in num_list:
    #Skips first row
    row_label = sheet.cell(row = rowNum+1,column = 1).value = rowNum

for colNum in num_list:
    #Skips column A
    column_label = sheet.cell(row = 1,column = colNum+1).value = colNum

# Populating Calculation Data 
for rowNum in range(1,sheet.max_row):

    for colNum in range(1,sheet.max_column):
        row_data = sheet.cell(row = rowNum+1,column = colNum+1).value = colNum*rowNum

wb.save('Multiplication_Table.xlsx')
print("File Multiplication_Table.xlsx generated")
