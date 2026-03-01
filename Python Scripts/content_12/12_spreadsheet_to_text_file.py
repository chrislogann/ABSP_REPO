from dotenv import load_dotenv
import os
import openpyxl
import re

"""
Script 12_spreadsheet_to_text_file.py takes excel workbook columns and transforms to text files.
"""

def write_text(pFolderPath):
    ## Open workbook
    os.chdir(pFolderPath)
    wb = openpyxl.open('text_file_lines.xlsx')
    sheet = wb.active

    ## Gather column data
    for colNum in range(sheet.max_column):

        ## Export each column as a text file.
        for rowNum in range(sheet.max_row):
            
            row_data = sheet.cell(row = rowNum+1,column = colNum+1).value

            if not row_data:
                continue

            if rowNum == 0:
            
                filename = str(row_data)
                pattern = r"^(.+?)\.(.+?)$"
                match = re.match(pattern,filename)
                
                filename = f"export_{match.group(1)}.txt"
                file = open(filename,'w')
                continue
            
            file.write(row_data)

        file.close()
        

load_dotenv()
input_folderpath = os.getenv("12_input_filepath")
write_text(input_folderpath)







