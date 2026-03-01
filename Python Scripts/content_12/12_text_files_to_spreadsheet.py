from dotenv import load_dotenv
import os
import openpyxl

"""
Script 12_text_files_to_spreadsheet reads a list of files
Each file populates a column and fills the rows below it.
A workbook is exported upon completion.
"""

def insert_workbook_data(pFileLine):
    ##Insert each file line into a column row

    workbook_filepath = os.path.join(os.getcwd(),'text_file_lines.xlsx')

    if os.path.exists(workbook_filepath):
        wb = openpyxl.open('text_file_lines.xlsx')
    elif not os.path.exists(workbook_filepath):
        wb = openpyxl.Workbook()
    
    wb.sheetnames
    sheet = wb.active
    sheet.title = 'Text File Lines'

    vFileLine_len = len(pFileLine)
    
    first_cell = sheet.cell(row = 1,column = 1).value

    if not first_cell:
        vColumn = sheet.max_column
    else:
        vColumn = sheet.max_column + 1

    for i in range(vFileLine_len):
        row_data = sheet.cell(row = i+1,column = vColumn).value = pFileLine[i]

    ##export excel workbook
    wb.save('text_file_lines.xlsx')

def get_text_lines(pDirectory):
    ##Gather files from folder
    os.chdir(pDirectory)

    contents = os.listdir(os.getcwd())

    for content in contents:

        if content.upper().endswith(".TXT"):

            file = open(content,"r")
            file_lines = file.readlines()
            file_lines.insert(0,content+"\n")
            insert_workbook_data(file_lines)
    
load_dotenv()
input_filepath = os.getenv("12_input_filepath")
get_text_lines(input_filepath)

